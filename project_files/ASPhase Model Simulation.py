import os
import sys
sys.path.insert(1, "../")

import networkx as nx
import ndlib.models.ModelConfig as mc
from ndlib.utils import multi_runs
import time
from ndlib.models.epidemics.ASPhaseModel import ASPhaseModel as p0
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook

if os.getcwd() != "/home/theakashain/Documents/Co-Op/COVID-19 Research/NDLib-C19-UNB/project_files/":
    os.chdir("/home/theakashain/Documents/Co-Op/COVID-19 Research/NDLib-C19-UNB/project_files/")
filepath = 'Test/Output4.xlsx'
#Create XLSX file
#wb = Workbook()
#ws = wb.active
#wb.save(filepath)
   
#Load XLSX file
wb = load_workbook(filename = filepath)
ws = wb.active

for cross_num in [0]:
    #Start timer
    start_time = time.time()
    #Population
    N = 20000
    #General connections
    connections = 12
    #Total number of days
    iterations = 30
    #Total number of tests
    executions = 1000
    #Initial infected
    initialinfect = 1
    #What day the test is at
    days = 7
    
    print('-----Generating Network With {} Nodes and {} Connections at Travel {}-----'.format(N, connections, cross_num))
    import NetworkGenerator as ng
    filename  = "connection_network.txt"
    filename2 = "age_network.txt"
    ng.NetworkGenerator(filename, filename2, N)
    G = nx.read_edgelist(filename)
    ages = []
    with open(filename2, "r") as age:
        for line in age:
                ages.append(float(line.split()[1]))
    #G = nx.watts_strogatz_graph(N, connections, p=0.01)
    
    # Model Selection
    print('-----Configuring Model-----')
    model = p0(G)
    
    #-----------------------------------------------------
    #DISEASE DATA
    r0 = 3.1
    disease_length = 14
    people_total = connections * disease_length 
    infect_chance = r0 / people_total
    #print(infect_chance)
    #-----------------------------------------------------
    # Model Configuration
    cfg = mc.Configuration()
    cfg.add_model_parameter('beta', infect_chance) #Probability to become infected by each neighbour
    cfg.add_model_parameter('gamma', (1 - 0.5**(1/5.8))) #Probability of recovery each day
    cfg.add_model_parameter('alpha', (1 - 0.5**(1/4.3))) #Probability of E -> Pre-symp each day
    cfg.add_model_parameter('phase', 0) #Starting phase (0 = yellow, 1 = orange, 2 = red)
    cfg.add_model_parameter('ptrans1', 6) #Infected people required to switch to orange
    cfg.add_model_parameter('ptrans2', 20) #Infected people required to switch to red
    cfg.add_model_parameter('asymp_chance', 0.20) #Chance of being asymptomatic
    cfg.add_model_parameter('asymp_infect', 1) #Chance of asymptomatic case causing infection
    cfg.add_model_parameter('symp_test', 0.1) #Chance of symptomatic case getting tested
    cfg.add_model_parameter('rand_test_count', 20) #Number of random tests per day
    cfg.add_model_parameter('days', days) #Day patient 0 is quarantined
    cfg.add_model_parameter("pre-symp_time", 2)
    cfg.add_model_parameter("test_chance", 0.9)
    cfg.add_model_parameter("crossing_num", cross_num)
    cfg.add_model_parameter("importation", 0.0016)
    
    #Give each edge a property of only yellow, only orange and yellow, or all phases
    for i in G.edges():
        ranvar = np.random.random()
        if (ranvar < 0.5): #Only Yellow
            cfg.add_edge_configuration("Phase Requirement", i, 0)
        elif (ranvar < 0.66): #Yellow and Orange
            cfg.add_edge_configuration("Phase Requirement",i, 0.5)
        else: #All Phases
            cfg.add_edge_configuration("Phase Requirement", i, 1)
            
    for i in G.nodes():
        cfg.add_node_configuration("Age", i, ages[int(i)])
    
    cfg.add_model_parameter('fraction_infected', initialinfect/N) #Set initial infected
    model.set_initial_status(cfg) #Input configuration above into model
    
    # Simulation
    print('-----Doing {} simulation(s) at day {}-----'.format(executions, days))
    trends = multi_runs(model, execution_number = executions, iteration_number = iterations, infection_sets=None)
    
    #End timer, set total time
    stop_time = time.time()
    total_time = stop_time - start_time
    print('----- Total Time: {} seconds ----'.format(total_time))
    
    print('----- Saving Results-----\n')
    
    #Store data from testing data in "daydata" array
    daydata = []  #Total Infected
    for n in range(0, executions):
        if (trends[n]['trends']['node_count'][0][-1] != (N - initialinfect)):
            daydata.append(N-trends[n]['trends']['node_count'][0][-1])
    infectdata = [] #Active Infected at last day
    for n in range(0, executions):
        if (trends[n]['trends']['node_count'][0][-1] != (N - initialinfect)):
            infectdata.append(trends[n]['trends']['node_count'][2][-1] +
                              trends[n]['trends']['node_count'][3][-1] +
                              trends[n]['trends']['node_count'][1][-1] + 
                              trends[n]['trends']['node_count'][8][-1])
    testeddata = [] #Tested at last day
    for n in range(0,executions):
        if (trends[n]['trends']['node_count'][0][-1] != (N - initialinfect)):
            testeddata.append(trends[n]['trends']['node_count'][5][-1] +
                              trends[n]['trends']['node_count'][6][-1])
    
    #Save data in XLSX file
    dayscol = (cross_num/250)*5+1
    totalcol = (cross_num/250)*5+2
    infectcol = (cross_num/250)*5+3
    testedcol = (cross_num/250)*5+4
    
    ws.cell(1, dayscol, "Crossing Num")
    ws.cell(1, totalcol, "Total Infected")
    ws.cell(1, infectcol, "Number Infected at Day 30")
    ws.cell(1, testedcol, "Number Tested at Day 30")
    
    row = 2
    for item in range(0,len(daydata)):
        ws.cell(row, dayscol, cross_num)
        row += 1
    row = 2
    for item in daydata:
        ws.cell(row, totalcol, item)
        row += 1
    row = 2
    for item in infectdata:
        ws.cell(row, infectcol, item)
        row += 1
    row = 2
    for item in testeddata:
        ws.cell(row, testedcol, item)
        row += 1

    wb.save(filepath)
    
    #for n in range(0, executions):
    #    print(trends[n])
