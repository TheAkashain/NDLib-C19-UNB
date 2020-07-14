import networkx as nx
import ndlib.models.ModelConfig as mc
from ndlib.utils import multi_runs
import matplotlib.pyplot as plt
import time
import ndlib.models.epidemics.ASPhaseModel as p0
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
import NetworkGenerator as ng

wb = Workbook()
ws = wb.active
wb.save('Test/Output3.xlsx')

#Choose day patient 0 is quarantined

for days in range(0,15):
    
    #Create XLSX file for data
    wb = load_workbook(filename = 'Test/Output3.xlsx')
    ws = wb.active
    
    #Start timer
    start_time = time.time()
    #Population
    N = 20000
    #General connections
    connections = 12
    #Total number of days
    iterations = 30
    #Total number of tests
    executions = 1000
    #Initial infected
    initialinfect = 1
    
    print('-----Generating network with {} nodes and {} connections-----'.format(N, connections))
    #filename  = "network1.txt"
    #ng.NetworkGenerator(filename, N)
    #g = nx.read_edgelist(filename)
    g= nx.barabasi_albert_graph(N, connections)
    
    # Model Selection
    print('-----Configuring Model-----')
    model = p0(g)
    
    #-----------------------------------------------------
    #DISEASE DATA
    r0 = 3.1
    disease_length = 14
    people_total = 12 * disease_length 
    chance_of_infection = r0 / people_total
    infect_chance = chance_of_infection
    #-----------------------------------------------------
    # Model Configuration
    cfg = mc.Configuration()
    cfg.add_model_parameter('beta', infect_chance) #Probability to become infected by each neighbour
    cfg.add_model_parameter('gamma', (1 - 0.5**(1/5.8))) #Probability of recovery each day
    cfg.add_model_parameter('alpha', (1 - 0.5**(1/4.3))) #Probability of E -> I each day
    cfg.add_model_parameter('phase', 0) #Starting phase (0 = yellow, 1 = orange, 2 = red)
    cfg.add_model_parameter('ptrans1', 6) #Infected people required to switch to orange
    cfg.add_model_parameter('ptrans2', 20) #Infected people required to switch to red
    cfg.add_model_parameter('asymp_chance', 0.25)
    cfg.add_model_parameter('asymp_infect', 1)
    cfg.add_model_parameter('symp_test', 0.1)
    cfg.add_model_parameter('rand_test_count', 20)
    
    #Give each edge a property of only yellow, only orange and yellow, or all phases
    for i in g.edges():
        ranvar = np.random.random()
        if (ranvar < 0.5):
            cfg.add_edge_configuration("Phase Requirement", i, 0)
        elif (ranvar < 0.66):
            cfg.add_edge_configuration("Phase Requirement",i, 0.5)
        else:
            cfg.add_edge_configuration("Phase Requirement", i, 1)
    
    cfg.add_model_parameter('fraction_infected', initialinfect/N) #Set initial infected
    model.set_initial_status(cfg) #Input configuration above into model
    
    # Simulation
    print('-----Doing {} simulation(s)-----'.format(executions))
    trends = multi_runs(model, execution_number = executions, iteration_number = iterations, infection_sets=None)
    
    #End timer, set total time
    stop_time = time.time()
    total_time = stop_time - start_time
    print('\n----- Total Time: {} seconds ----'.format(total_time))
    
    print('-----Plotting Results-----')
    
    #Store data from testing data in "daydata" array
    daydata = []  
    for n in range(0, executions):
        if (trends[n]['trends']['node_count'][0][-1] != (N - initialinfect)):
            daydata.append(N-trends[n]['trends']['node_count'][0][-1])
    
    #Set up and do figures
    fig = plt.hist(daydata)
    stdev = np.std(daydata)
    mean = np.mean(daydata)
    
    #Save data in XLSX file
    row = 2
    dayscol = days*3+1
    column = (days*3)+2
    ws.cell(1, dayscol, "Day Quarantined")
    ws.cell(1, column, "Number Infected")
    for item in daydata:
        ws.cell(row, dayscol,days)
        ws.cell(row,column,item)
        row += 1
    
    #Finish plotting
    plt.title("Infected at Day {}".format(iterations), fontsize = 18)
    plt.xlabel("Infected Population", fontsize = 18)
    plt.ylabel("Number of Simulations", fontsize = 18)
    plt.figtext(0.55, 0.70, 'Outbreaks: {:.0f}\nMean: {:.3f}\nStDev: {:.3f}\nExecution Time: {:.3f}'.format(len(daydata), mean, stdev, total_time), 
            fontsize = 15, bbox = dict(boxstyle = 'round', facecolor = 'white'))
    plt.tight_layout()
    plt.savefig("Test/(HIST)Patient 0 Test: {} People, {} Days Test, {} Days Total, {} Initial.png".format(N, days, iterations, initialinfect))
    plt.clf()
    plt.close()
    wb.save('Test/Output3.xlsx')